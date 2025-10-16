# -*- coding: utf-8 -*-
"""
Module xuất báo cáo Excel và PDF
Report Export Module for Attendance System
"""

import os
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, Reference
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.units import inch
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import config
from database import DatabaseManager

class ReportExporter:
    """Class xuất báo cáo chấm công"""
    
    def __init__(self):
        self.db = DatabaseManager()
        
        # Register font hỗ trợ tiếng Việt cho PDF (nếu có)
        try:
            # Thử load font hỗ trợ Unicode
            # pdfmetrics.registerFont(TTFont('TimesNewRoman', 'times.ttf'))
            pass
        except:
            print("⚠️  Could not load Unicode font for PDF. Using default font.")
    
    def export_to_excel(self, start_date=None, end_date=None, employee_id=None, output_path=None):
        """
        Xuất báo cáo chấm công ra file Excel với formatting đẹp
        
        Args:
            start_date: Ngày bắt đầu (format: DD/MM/YYYY)
            end_date: Ngày kết thúc (format: DD/MM/YYYY)
            employee_id: ID nhân viên (None = tất cả)
            output_path: Đường dẫn file output
        
        Returns:
            output_path: Đường dẫn file đã tạo
        """
        print("📊 Generating Excel report...")
        
        # Lấy dữ liệu
        logs = self.db.get_attendance_logs(start_date, end_date, employee_id)
        
        if not logs:
            print("❌ No data to export!")
            return None
        
        # Tạo DataFrame
        df = pd.DataFrame(logs)
        
        # Format output path
        if output_path is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"BaoCaoChamCong_{timestamp}.xlsx"
            output_path = os.path.join(config.REPORTS_DIR, filename)
        
        # Tạo Excel Writer
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Sheet 1: Chi tiết chấm công
            df_detail = df[['datetime', 'employee_id', 'employee_name', 'department', 
                           'type', 'status', 'confidence', 'is_late', 'notes']]
            df_detail.columns = ['Thời gian', 'Mã NV', 'Họ tên', 'Phòng ban', 
                                'Loại', 'Trạng thái', 'Độ tin cậy', 'Đi muộn', 'Ghi chú']
            df_detail.to_excel(writer, sheet_name='Chi tiết', index=False)
            
            # Sheet 2: Thống kê theo nhân viên
            summary = df.groupby(['employee_id', 'employee_name', 'department']).agg({
                'id': 'count',
                'is_late': 'sum',
                'confidence': 'mean'
            }).reset_index()
            summary.columns = ['Mã NV', 'Họ tên', 'Phòng ban', 'Số lần chấm', 'Số lần muộn', 'Độ tin cậy TB']
            summary.to_excel(writer, sheet_name='Thống kê NV', index=False)
            
            # Sheet 3: Thống kê theo ngày
            df['date'] = pd.to_datetime(df['datetime'], format=config.DATETIME_FORMAT).dt.date
            daily_summary = df.groupby('date').agg({
                'id': 'count',
                'is_late': 'sum'
            }).reset_index()
            daily_summary.columns = ['Ngày', 'Tổng lượt chấm', 'Số lần muộn']
            daily_summary.to_excel(writer, sheet_name='Thống kê theo ngày', index=False)
            
            # Sheet 4: Thống kê theo phòng ban
            dept_summary = df.groupby('department').agg({
                'id': 'count',
                'is_late': 'sum',
                'confidence': 'mean'
            }).reset_index()
            dept_summary.columns = ['Phòng ban', 'Số lần chấm', 'Số lần muộn', 'Độ tin cậy TB']
            dept_summary.to_excel(writer, sheet_name='Thống kê phòng ban', index=False)
        
        # Format Excel file
        self._format_excel(output_path)
        
        print(f"✅ Excel report exported: {output_path}")
        return output_path
    
    def _format_excel(self, file_path):
        """Format Excel file với styles đẹp"""
        wb = load_workbook(file_path)
        
        # Define styles
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Format header row
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            # Format data rows
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # Auto-adjust column width
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Freeze first row
            ws.freeze_panes = 'A2'
        
        wb.save(file_path)
    
    def export_to_pdf(self, start_date=None, end_date=None, employee_id=None, output_path=None):
        """
        Xuất báo cáo chấm công ra file PDF
        
        Args:
            start_date: Ngày bắt đầu (format: DD/MM/YYYY)
            end_date: Ngày kết thúc (format: DD/MM/YYYY)
            employee_id: ID nhân viên (None = tất cả)
            output_path: Đường dẫn file output
        
        Returns:
            output_path: Đường dẫn file đã tạo
        """
        print("📄 Generating PDF report...")
        
        # Lấy dữ liệu
        logs = self.db.get_attendance_logs(start_date, end_date, employee_id)
        
        if not logs:
            print("❌ No data to export!")
            return None
        
        # Format output path
        if output_path is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"BaoCaoChamCong_{timestamp}.pdf"
            output_path = os.path.join(config.REPORTS_DIR, filename)
        
        # Tạo PDF
        doc = SimpleDocTemplate(output_path, pagesize=landscape(A4))
        elements = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#1F4E78'),
            spaceAfter=30,
            alignment=1  # Center
        )
        
        title = Paragraph(f"<b>{config.PDF_TITLE}</b>", title_style)
        elements.append(title)
        
        # Company name
        if config.COMPANY_NAME:
            company = Paragraph(f"<b>{config.COMPANY_NAME}</b>", styles['Normal'])
            elements.append(company)
            elements.append(Spacer(1, 12))
        
        # Date range
        period = f"Từ ngày: {start_date or 'Đầu'} - Đến ngày: {end_date or 'Hiện tại'}"
        elements.append(Paragraph(period, styles['Normal']))
        elements.append(Spacer(1, 20))
        
        # Summary statistics
        total_records = len(logs)
        total_late = sum(1 for log in logs if log['is_late'])
        unique_employees = len(set(log['employee_id'] for log in logs))
        
        summary_text = f"""
        <b>Tổng quan:</b><br/>
        - Tổng số lượt chấm công: {total_records}<br/>
        - Số nhân viên: {unique_employees}<br/>
        - Số lần đi muộn: {total_late}<br/>
        """
        elements.append(Paragraph(summary_text, styles['Normal']))
        elements.append(Spacer(1, 20))
        
        # Table data
        table_data = [['Thời gian', 'Mã NV', 'Họ tên', 'Phòng ban', 'Loại', 'Trạng thái', 'Đi muộn']]
        
        for log in logs[:100]:  # Limit to first 100 records for PDF
            table_data.append([
                log['datetime'],
                log['employee_id'],
                log['employee_name'],
                log['department'] or '',
                log['type'],
                log['status'],
                'Có' if log['is_late'] else 'Không'
            ])
        
        # Create table
        table = Table(table_data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E78')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
        ]))
        
        elements.append(table)
        
        if len(logs) > 100:
            elements.append(Spacer(1, 12))
            note = Paragraph(f"<i>* Chỉ hiển thị 100/{len(logs)} bản ghi đầu tiên</i>", styles['Normal'])
            elements.append(note)
        
        # Build PDF
        doc.build(elements)
        
        print(f"✅ PDF report exported: {output_path}")
        return output_path
    
    def generate_full_report(self, start_date=None, end_date=None, employee_id=None):
        """
        Tạo báo cáo đầy đủ (cả Excel và PDF)
        
        Returns:
            excel_path, pdf_path: Tuple đường dẫn 2 file
        """
        print("📊 Generating full report (Excel + PDF)...")
        
        excel_path = self.export_to_excel(start_date, end_date, employee_id)
        pdf_path = self.export_to_pdf(start_date, end_date, employee_id)
        
        return excel_path, pdf_path


# Test
if __name__ == "__main__":
    exporter = ReportExporter()
    
    # Test export
    print("\n🧪 Testing report export...")
    
    # Uncomment to test
    # excel_file = exporter.export_to_excel()
    # pdf_file = exporter.export_to_pdf()
    
    print("✅ Test completed!")
